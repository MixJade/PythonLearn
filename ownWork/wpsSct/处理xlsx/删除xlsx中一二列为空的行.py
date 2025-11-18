# coding=utf-8
# @Time    : 2025/7/21 9:35
# @Software: PyCharm
import openpyxl


def delete_rows_with_empty_first_two_columns(file_path):
    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)

    # 遍历所有表
    for sheet in workbook.worksheets:
        # 从最后一行开始向前遍历，避免删除行后索引混乱
        for row_idx in range(sheet.max_row, 0, -1):
            # 获取第一列和第二列的值
            cell_a = sheet.cell(row=row_idx, column=1).value
            cell_b = sheet.cell(row=row_idx, column=2).value

            # 如果第一列和第二列都为空，则删除该行
            if cell_a is None and cell_b is None:
                sheet.delete_rows(row_idx)

    # 保存修改后的工作簿
    output_file = file_path.replace('.xlsx', '_modified.xlsx')
    workbook.save(output_file)
    print(f"处理完成，已保存至: {output_file}")


# 使用示例
if __name__ == "__main__":
    file_path_1 = "your_excel_file.xlsx"  # 替换为实际文件路径
    delete_rows_with_empty_first_two_columns(file_path_1)
