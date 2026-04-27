# coding=utf-8
# @Time    : 2026/4/27 16:26
# @Software: PyCharm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 1. 打开你的 Excel 文件（替换成你的文件路径）
wb = load_workbook(r"3年前领补贴的同事名单.xlsx")

# 2. 设置要操作的工作表（默认第一个工作表，也可以写名字：wb["Sheet1"]）
ws = wb.active

# 3. 定义黄色填充样式
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 4. 遍历所有有数据的单元格
for row in ws.iter_rows():  # 遍历每一行
    for cell in row:  # 遍历每一行里的每个单元格
        # 判断单元格是否有内容，并且内容包含“离职”
        if cell.value is not None and "离职" in str(cell.value):
            cell.fill = yellow_fill  # 设置黄色背景

# 5. 保存文件（可以另存为新文件，避免覆盖原数据）
wb.save("标黄内容_结果.xlsx")
print("处理完成！文件已保存为：标黄内容_结果.xlsx")
