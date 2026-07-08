# coding=utf-8
# @Time    : 2026/7/7 20:45
# @Software: PyCharm
"""
CSV 合并为 XLSX
通过 input 依次输入 CSV 文件路径，输入 0 停止，
将所有 CSV 合并到一个 xlsx 文件中，输出到桌面。
每个 sheet 页名称为对应 CSV 的文件名（无后缀）。
"""

import os
import csv
import openpyxl
from openpyxl.utils import get_column_letter


def get_desktop_path():
    """获取桌面路径"""
    return os.path.join(os.path.expanduser("~"), "Desktop")


def collect_csv_paths():
    """通过 input 收集 CSV 文件路径，输入 0 结束"""
    paths = []
    print("请输入 CSV 文件路径（输完后输入 0 结束）：")
    idx = 1
    while True:
        path = input(f"第 {idx} 个: ").strip().strip('"')
        if path == "0":
            print("已停止输入。")
            break
        if not path:
            print("路径不能为空，请重新输入。")
            continue
        if not os.path.isfile(path):
            print(f"文件不存在: {path}")
            continue
        if not path.lower().endswith(".csv"):
            print(f"不是 CSV 文件: {path}")
            continue
        paths.append(path)
        print(f"  -> 已添加: {path}")
        idx += 1
    return paths


def csv_to_xlsx(csv_paths, output_path):
    """将所有 CSV 文件写入同一个 xlsx 的不同 sheet"""
    wb = openpyxl.Workbook()
    # 删除默认的 Sheet
    wb.remove(wb.active)

    for csv_path in csv_paths:
        # sheet 名 = 文件名（不含后缀），截断到 31 字符（Excel 限制）
        sheet_name = os.path.splitext(os.path.basename(csv_path))[0]
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        print(f"正在处理: {sheet_name} ...")

        ws = wb.create_sheet(title=sheet_name)

        with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # 自适应列宽（简单处理）
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells:
                if cell.value:
                    # 中文字符按 2 倍宽度估算
                    char_len = 0
                    for ch in str(cell.value):
                        char_len += 2 if ord(ch) > 127 else 1
                    max_len = max(max_len, char_len)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(output_path)
    print(f"\n完成！已输出到: {output_path}")


def main():
    csv_paths = collect_csv_paths()
    if not csv_paths:
        print("未添加任何 CSV 文件，退出。")
        return

    output = os.path.join(get_desktop_path(), "合并结果.xlsx")
    csv_to_xlsx(csv_paths, output)


if __name__ == "__main__":
    main()
